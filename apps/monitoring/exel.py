import openpyxl

from helper import Helper


class Excel:
    workbook = None
    sheet = None

    @staticmethod
    def open(file_path):
        Excel.workbook = openpyxl.load_workbook(file_path)
        Excel.sheet = Excel.workbook.active

    @staticmethod
    def save(file_path):
        Excel.workbook.save(file_path)

    @staticmethod
    def add_measurements(connected_sensors):
        existing_sensor_names = Excel.read_existing_sensor_names()
        not_existing_sensor_names = Excel.get_not_existing_sensor_names(
            connected_sensors, existing_sensor_names)
        Excel.add_not_existing_sensor_names(not_existing_sensor_names)
        existing_sensor_names = Excel.read_existing_sensor_names()
        data_row = Excel.generate_measurement_data_row(
            connected_sensors, existing_sensor_names)
        Excel.add_data_row_to_file(data_row)

    @staticmethod
    def read_existing_sensor_names():
        first_row = Excel.read_first_row()
        return first_row[2:]

    @staticmethod
    def get_not_existing_sensor_names(connected_sensors, existing_sensor_names):
        not_existing_sensor_names = []
        for sensor in connected_sensors:
            sensor_name = sensor.get_name()
            sensor_name_exists = False
            for existing_sensor_name in existing_sensor_names:
                if existing_sensor_name == sensor_name:
                    sensor_name_exists = True
                    break
            if not sensor_name_exists:
                not_existing_sensor_names.append(sensor_name)
        return not_existing_sensor_names

    @staticmethod
    def add_not_existing_sensor_names(not_existing_sensor_names):
        max_column = Excel.get_max_column()
        for i, sensor_name in enumerate(not_existing_sensor_names):
            column = max_column + i + 1
            cell = Excel.sheet.cell(row=1, column=column)
            cell.value = sensor_name

    @staticmethod
    def generate_measurement_data_row(connected_sensors, existing_sensor_names):
        data_row = [Helper.get_current_date_string(), Helper.get_current_time_string()]
        for i, existing_sensor_name in enumerate(existing_sensor_names):
            sensor_value = ''
            for connected_sensor in connected_sensors:
                if connected_sensor.get_name() == existing_sensor_name:
                    sensor_value = connected_sensor.get_value()
                    break
            data_row.append(sensor_value)
        return data_row

    @staticmethod
    def read_first_row():
        first_row = []
        for i in range(1, Excel.get_max_column() + 1):
            cell = Excel.sheet.cell(row=1, column=i)
            first_row.append(cell.value)
        return first_row

    @staticmethod
    def add_data_row_to_file(data_row):
        Excel.sheet.append(data_row)

    @staticmethod
    def get_max_column():
        return Excel.sheet.max_column

    @staticmethod
    def get_max_row():
        return Excel.sheet.max_row
